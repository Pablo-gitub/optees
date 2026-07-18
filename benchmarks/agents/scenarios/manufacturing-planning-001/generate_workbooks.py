from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
REFERENCE_DIR = ROOT / "reference"

NAVY = "101827"
BLUE = "1D4ED8"
TEAL = "14B8A6"
PALE_BLUE = "DBEAFE"
PALE_TEAL = "CCFBF1"
PALE_AMBER = "FEF3C7"
WHITE = "FFFFFF"
INK = "172033"
MUTED = "5B657A"
GRID = "CBD5E1"
DOCUMENT_TIMESTAMP = datetime(2026, 7, 18, 0, 0, 0)
ZIP_TIMESTAMP = (2026, 7, 18, 0, 0, 0)

HISTORY_ROWS = [
    ("2025-01", 3, 0.8, 12),
    ("2025-02", 5, 1.0, 11),
    ("2025-03", 2, 1.2, 13),
    ("2025-04", 7, 0.9, 12),
    ("2025-05", 4, 1.1, 10),
    ("2025-06", 6, 1.3, 14),
    ("2025-07", 8, 0.7, 13),
    ("2025-08", 1, 1.4, 11),
    ("2025-09", 9, 1.0, 15),
    ("2025-10", 5, 0.6, 9),
    ("2025-11", 3, 1.5, 14),
    ("2025-12", 7, 1.2, 10),
]


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REFERENCE_DIR.mkdir(parents=True, exist_ok=True)
    input_path = DATA_DIR / "fictional_company_input.xlsx"
    reference_path = REFERENCE_DIR / "fictional_company_ground_truth.xlsx"
    _build_input(input_path)
    _build_reference(reference_path)
    _validate(input_path, reference_path)
    print(input_path)
    print(reference_path)


def _build_input(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("Read Me")
    _title(readme, "A1", "Fictitious Manufacturing Company")
    readme["A3"] = "Scenario"
    readme["B3"] = "Northstar Components - next-period production planning"
    readme["A4"] = "Data status"
    readme["B4"] = "Fully synthetic; no real company or personal data"
    readme["A6"] = "Workbook use"
    readme["B6"] = (
        "Use Direct Demand Plan for the single-solver task. For the orchestration "
        "task, estimate demand from History A, History B, and Next Period before "
        "optimizing production."
    )
    readme["A8"] = "Business objective"
    readme["B8"] = (
        "Maximize total contribution while respecting product commitments, demand "
        "limits, and every hard resource capacity. Production quantities are whole units."
    )
    readme["A10"] = "Important"
    readme["B10"] = (
        "Administrative Notes contains contextual fields that are not mathematical "
        "constraints. Do not convert them into constraints unless the task explicitly says so."
    )
    _label_block(readme, "A3:A10")
    readme["B10"].fill = PatternFill("solid", fgColor=PALE_AMBER)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 105
    for row in (6, 8, 10):
        readme.row_dimensions[row].height = 42
    readme["B6"].alignment = readme["B8"].alignment = readme["B10"].alignment = Alignment(
        wrap_text=True, vertical="top"
    )

    products = wb.create_sheet("Products")
    _title(products, "A1", "Products and unit economics")
    product_rows = [
        [
            "Product ID",
            "Description",
            "Unit contribution EUR",
            "Machine hours / unit",
            "Material kg / unit",
            "Labor hours / unit",
            "Minimum units",
            "Whole units required",
            "Account manager",
            "Brand color",
        ],
        ["A", "Precision housing", 40, 2, 3, 1, 5, True, "Elena", "Blue"],
        ["B", "Control module", 55, 4, 2, 2, 3, True, "Marco", "Green"],
    ]
    _write_table(products, "A3:J5", product_rows, "ProductsTable")
    _number(products, "C4:G5", "#,##0.00")
    products.freeze_panes = "A4"
    _widths(products, [14, 25, 23, 22, 20, 20, 17, 21, 19, 15])

    resources = wb.create_sheet("Resources")
    _title(resources, "A1", "Hard production capacities")
    resource_rows = [
        ["Resource", "Available capacity", "Unit", "Hard constraint", "Planning owner"],
        ["Machine time", 60, "hours", True, "Operations"],
        ["Raw material", 80, "kg", True, "Procurement"],
        ["Direct labor", 35, "hours", True, "Operations"],
    ]
    _write_table(resources, "A3:E7", resource_rows, "ResourcesTable")
    _number(resources, "B4:B6", "#,##0.00")
    resources.freeze_panes = "A4"
    _widths(resources, [22, 21, 15, 19, 20])

    direct = wb.create_sheet("Direct Demand Plan")
    _title(direct, "A1", "Approved demand caps for the single-solver task")
    direct_rows = [
        ["Product ID", "Maximum demand units", "Approval status", "Source note"],
        ["A", 24, "Approved", "Commercial planning estimate"],
        ["B", 10, "Approved", "Commercial planning estimate"],
    ]
    _write_table(direct, "A3:D5", direct_rows, "DirectDemandTable")
    _number(direct, "B4:B5", "#,##0")
    _widths(direct, [15, 24, 20, 35])

    history_a = wb.create_sheet("History A")
    _history_sheet(history_a, "A", lambda m, s, p: 20 + 3 * m + 5 * s - 2 * p)
    history_b = wb.create_sheet("History B")
    _history_sheet(history_b, "B", lambda m, s, p: 10 + 2 * m + 4 * s - p)

    next_period = wb.create_sheet("Next Period")
    _title(next_period, "A1", "Known explanatory variables for next period")
    next_rows = [
        [
            "Product ID",
            "Marketing spend kEUR",
            "Season index",
            "Unit price EUR",
            "Demand forecast units",
            "Forecast status",
        ],
        ["A", 4, 1.2, 10, None, "To be estimated"],
        ["B", 7, 1.2, 15, None, "To be estimated"],
    ]
    _write_table(next_period, "A3:F5", next_rows, "NextPeriodTable")
    _number(next_period, "B4:E5", "#,##0.00")
    next_period["E4"].fill = next_period["E5"].fill = PatternFill(
        "solid", fgColor=PALE_AMBER
    )
    _widths(next_period, [15, 25, 17, 19, 24, 22])

    notes = wb.create_sheet("Administrative Notes")
    _title(notes, "A1", "Contextual fields - not optimization constraints")
    note_rows = [
        ["Field", "Value", "Use in current tasks"],
        ["Warehouse code", "NSC-01", "Identification only"],
        ["Presentation currency", "EUR", "Report formatting only"],
        ["Preferred report language", "English", "Report formatting only"],
        ["Corporate color", "Navy", "Report styling only"],
        ["Energy review", "Scheduled next quarter", "No numeric limit supplied"],
    ]
    _write_table(notes, "A3:C8", note_rows, "AdministrativeTable")
    _widths(notes, [28, 30, 38])

    dictionary = wb.create_sheet("Data Dictionary")
    _title(dictionary, "A1", "Field definitions and units")
    dictionary_rows = [
        ["Field", "Definition", "Unit / domain"],
        ["Unit contribution", "Revenue less variable cost per produced unit", "EUR / unit"],
        ["Minimum units", "Mandatory minimum production commitment", "whole units"],
        ["Maximum demand", "Largest quantity that can be sold in the period", "whole units"],
        ["Marketing spend", "Planned commercial spend", "thousand EUR"],
        ["Season index", "Synthetic seasonal demand indicator", "dimensionless"],
        ["Observed demand", "Historical units requested by customers", "units"],
    ]
    _write_table(dictionary, "A3:C9", dictionary_rows, "DictionaryTable")
    _widths(dictionary, [25, 65, 24])

    for sheet in wb.worksheets:
        _finish_sheet(sheet)
    _save_deterministic(wb, path)


def _build_reference(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    guide = wb.create_sheet("Evaluation Guide")
    _title(guide, "A1", "Private benchmark ground truth - do not share with the agent")
    guide["A3"] = "Single-solver expected capability"
    guide["B3"] = "milp.linear"
    guide["A4"] = "Orchestration expected sequence"
    guide["B4"] = "ml.regression.linear twice, then milp.linear"
    guide["A6"] = "Invalidating errors"
    guide["B6"] = (
        "Using LP despite whole-unit requirements; treating administrative fields as hard "
        "constraints; using Direct Demand Plan in the orchestration task; silently rounding "
        "forecasts without explaining integer upper-bound semantics."
    )
    _label_block(guide, "A3:A6")
    guide.column_dimensions["A"].width = 38
    guide.column_dimensions["B"].width = 105
    guide["B6"].alignment = Alignment(wrap_text=True, vertical="top")
    guide.row_dimensions[6].height = 55

    single = wb.create_sheet("Single Solver Reference")
    _title(single, "A1", "Expected direct-demand optimization result")
    single_rows = [
        ["Metric", "Product A", "Product B", "Total / status"],
        ["Production units", 24, 3, 27],
        ["Contribution EUR", 960, 165, 1125],
        ["Machine hours", 48, 12, 60],
        ["Material kg", 72, 6, 78],
        ["Labor hours", 24, 6, 30],
        ["Demand upper bound", 24, 10, "A binding"],
        ["Minimum production", 5, 3, "B binding"],
    ]
    _write_table(single, "A3:D11", single_rows, "SingleReferenceTable")
    _number(single, "B4:C11", "#,##0.00")
    _widths(single, [28, 18, 18, 22])

    reg_a = wb.create_sheet("Regression A Reference")
    _regression_reference(reg_a, "A", 20, 3, 5, -2, 18)
    reg_b = wb.create_sheet("Regression B Reference")
    _regression_reference(reg_b, "B", 10, 2, 4, -1, 13.8)

    orchestration = wb.create_sheet("Orchestration Reference")
    _title(orchestration, "A1", "Expected forecast-to-production result")
    orchestration_rows = [
        ["Metric", "Product A", "Product B", "Total / status"],
        ["Raw forecast", 18, 13.8, "Regression output"],
        ["Largest feasible integer under forecast", 18, 13, "Bounds for MILP"],
        ["Production units", 18, 6, 24],
        ["Contribution EUR", 720, 330, 1050],
        ["Machine hours", 36, 24, 60],
        ["Material kg", 54, 12, 66],
        ["Labor hours", 18, 12, 30],
        ["Binding constraints", "Demand cap", "None", "Machine time"],
    ]
    _write_table(
        orchestration,
        "A3:D12",
        orchestration_rows,
        "OrchestrationReferenceTable",
    )
    _number(orchestration, "B4:C11", "#,##0.00")
    _widths(orchestration, [38, 20, 20, 25])

    rubric = wb.create_sheet("Rubric")
    _title(rubric, "A1", "Agent evaluation rubric")
    rubric_rows = [
        ["Criterion", "Maximum points", "Full-credit evidence"],
        ["Workbook data extraction", 15, "Uses relevant cells with correct units"],
        ["Capability selection", 15, "Selects expected solver sequence"],
        ["Payload validity", 15, "Validates each exact versioned payload"],
        ["Numerical correctness", 20, "Matches reference forecasts and optimum"],
        ["Status discipline", 10, "Separates solver and independent validation status"],
        ["Assumption discipline", 10, "No silent assumptions or invented data"],
        ["Report traceability", 10, "DOCX/PDF values trace to workbook and Optees"],
        ["Limitations", 5, "States forecast and modeling limitations"],
        ["Total", 100, ""],
    ]
    _write_table(rubric, "A3:C12", rubric_rows, "RubricTable")
    _widths(rubric, [32, 20, 72])

    for sheet in wb.worksheets:
        _finish_sheet(sheet)
    _save_deterministic(wb, path)


def _history_sheet(sheet, product: str, demand_formula) -> None:
    _title(sheet, "A1", f"Synthetic historical demand - Product {product}")
    rows = [[
        "Period",
        "Marketing spend kEUR",
        "Season index",
        "Unit price EUR",
        "Observed demand units",
    ]]
    for period, marketing, season, price in HISTORY_ROWS:
        rows.append(
            [period, marketing, season, price, demand_formula(marketing, season, price)]
        )
    _write_table(sheet, "A3:E15", rows, f"History{product}Table")
    _number(sheet, "B4:E15", "#,##0.00")
    sheet.freeze_panes = "A4"
    _widths(sheet, [15, 25, 17, 19, 25])


def _regression_reference(
    sheet,
    product: str,
    intercept: float,
    marketing: float,
    season: float,
    price: float,
    forecast: float,
) -> None:
    _title(sheet, "A1", f"Expected OLS model - Product {product}")
    rows = [
        ["Term", "Expected coefficient", "Interpretation"],
        ["Intercept", intercept, "Baseline synthetic demand"],
        ["Marketing spend kEUR", marketing, "Marginal units per kEUR"],
        ["Season index", season, "Seasonality coefficient"],
        ["Unit price EUR", price, "Price coefficient"],
        ["Next-period forecast", forecast, "Model prediction before integer planning"],
    ]
    _write_table(sheet, "A3:C9", rows, f"Regression{product}ReferenceTable")
    _number(sheet, "B4:B9", "#,##0.0000")
    _widths(sheet, [28, 24, 48])


def _title(sheet, cell: str, text: str) -> None:
    sheet[cell] = text
    sheet[cell].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    sheet[cell].fill = PatternFill("solid", fgColor=NAVY)
    sheet[cell].alignment = Alignment(vertical="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    sheet.row_dimensions[1].height = 34


def _write_table(sheet, range_ref: str, rows: list[list[object]], name: str) -> None:
    start, _end = range_ref.split(":")
    anchor = sheet[start]
    for row_offset, values in enumerate(rows):
        for col_offset, value in enumerate(values):
            sheet.cell(
                row=anchor.row + row_offset,
                column=anchor.column + col_offset,
                value=value,
            )
    actual_end_row = anchor.row + len(rows) - 1
    actual_end_col = anchor.column + len(rows[0]) - 1
    actual_ref = (
        f"{anchor.coordinate}:"
        f"{sheet.cell(row=actual_end_row, column=actual_end_col).coordinate}"
    )
    table = Table(displayName=name, ref=actual_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _finish_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = sheet.freeze_panes or "A3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_view.zoomScale = 90
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row > 1:
                font = copy(cell.font)
                font.name = "Aptos"
                font.color = INK
                cell.font = font
                if cell.value is not None:
                    alignment = copy(cell.alignment)
                    alignment.vertical = "center"
                    cell.alignment = alignment


def _label_block(sheet, range_ref: str) -> None:
    thin = Side(style="thin", color=GRID)
    for row in sheet[range_ref]:
        for cell in row:
            cell.font = Font(name="Aptos", bold=True, color=INK)
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
            cell.border = Border(bottom=thin)


def _number(sheet, range_ref: str, number_format: str) -> None:
    for row in sheet[range_ref]:
        for cell in row:
            cell.number_format = number_format


def _widths(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _save_deterministic(workbook: Workbook, path: Path) -> None:
    workbook.properties.creator = "Optees"
    workbook.properties.lastModifiedBy = "Optees"
    workbook.properties.created = DOCUMENT_TIMESTAMP
    workbook.properties.modified = DOCUMENT_TIMESTAMP
    workbook.save(path)

    normalized_path = path.with_suffix(".normalized.xlsx")
    with ZipFile(path, "r") as source, ZipFile(
        normalized_path,
        "w",
        compression=ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for source_info in sorted(source.infolist(), key=lambda item: item.filename):
            target_info = ZipInfo(source_info.filename, ZIP_TIMESTAMP)
            target_info.compress_type = (
                ZIP_STORED if source_info.is_dir() else ZIP_DEFLATED
            )
            target_info.create_system = source_info.create_system
            target_info.external_attr = source_info.external_attr
            target_info.internal_attr = source_info.internal_attr
            target_info.flag_bits = source_info.flag_bits
            target.writestr(target_info, source.read(source_info.filename))
    normalized_path.replace(path)


def _validate(input_path: Path, reference_path: Path) -> None:
    expected_input = {
        "Read Me",
        "Products",
        "Resources",
        "Direct Demand Plan",
        "History A",
        "History B",
        "Next Period",
        "Administrative Notes",
        "Data Dictionary",
    }
    expected_reference = {
        "Evaluation Guide",
        "Single Solver Reference",
        "Regression A Reference",
        "Regression B Reference",
        "Orchestration Reference",
        "Rubric",
    }
    input_wb = load_workbook(input_path, data_only=False)
    reference_wb = load_workbook(reference_path, data_only=False)
    assert set(input_wb.sheetnames) == expected_input
    assert set(reference_wb.sheetnames) == expected_reference
    assert input_wb["Products"]["C4"].value == 40
    assert input_wb["Next Period"]["E4"].value is None
    assert reference_wb["Single Solver Reference"]["D5"].value == 1125
    assert reference_wb["Orchestration Reference"]["D7"].value == 1050


if __name__ == "__main__":
    main()
